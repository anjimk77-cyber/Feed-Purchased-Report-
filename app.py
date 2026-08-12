"""
Customer Feed Purchase Report
------------------------------
Reads sales data from a Google Sheet, maps customers to Zones using an
uploaded Customer List (Excel: Customer ID, Zone, ...), and shows a
report with:
    Customer Code | Customer Name | Last Feed Purchase Date |
    Due Date (Last Purchase + N days) | Last Order Date

Zone-wise slicer in the sidebar filters the whole report.

Run with:
    streamlit run app.py
"""

import io
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

st.set_page_config(page_title="Customer Feed Purchase Report", layout="wide")

# ----------------------------------------------------------------------
# CONFIG — edit these if your sheet changes
# ----------------------------------------------------------------------
SHEET_ID = "1S3csAE-E_hN8vstuHR0KkeAN7yCVQTFe4AkEVlw4vQw"
DEFAULT_GID = "0"                 # tab (gid) of the sales data sheet
FEED_PREFIX = "FEED"              # Item No. prefix that identifies "feed" items
CUSTOMER_LIST_PATH = "Customer List.xlsx"   # <-- change to match the filename you committed to GitHub

# ----------------------------------------------------------------------
# DATA LOADERS
# ----------------------------------------------------------------------
@st.cache_data(ttl=300, show_spinner="Loading sales data from Google Sheet...")
def load_sales_data(sheet_id: str, gid: str) -> pd.DataFrame:
    """
    Loads the sales log from a published/shared Google Sheet.
    Requires sheet sharing: "Anyone with the link -> Viewer".
    """
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    df = pd.read_csv(url)

    df.columns = [c.strip() for c in df.columns]
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df["Customer Code"] = df["Customer Code"].astype(str).str.strip()
    df["Item No."] = df["Item No."].astype(str).str.strip()
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0)
    df["Sales Amt"] = pd.to_numeric(df["Sales Amt"], errors="coerce").fillna(0)
    return df


@st.cache_data(ttl=300, show_spinner="Reading customer / zone list...")
def load_customer_master(path: str) -> pd.DataFrame:
    """
    Reads the Customer List Excel file committed to your GitHub repo.
    Expected columns (case-insensitive, flexible naming):
        Customer ID / Customer Code   -> customer code
        Zone                          -> zone
        Customer Name (optional)      -> fallback name
    """
    raw = pd.read_excel(path)
    raw.columns = [c.strip() for c in raw.columns]

    # Flexible column matching
    col_map = {}
    for c in raw.columns:
        cl = c.lower()
        if cl in ("customer id", "customer code", "cust id", "cust code"):
            col_map[c] = "Customer Code"
        elif cl == "zone":
            col_map[c] = "Zone"
        elif cl in ("customer name", "cust name", "name"):
            col_map[c] = "Customer Name (Master)"

    raw = raw.rename(columns=col_map)

    if "Customer Code" not in raw.columns or "Zone" not in raw.columns:
        raise ValueError(
            "Could not find 'Customer ID/Code' and 'Zone' columns in "
            f"'{path}'. Found columns: " + ", ".join(raw.columns)
        )

    raw["Customer Code"] = raw["Customer Code"].astype(str).str.strip()
    raw["Zone"] = raw["Zone"].astype(str).str.strip()

    keep_cols = ["Customer Code", "Zone"]
    if "Customer Name (Master)" in raw.columns:
        keep_cols.append("Customer Name (Master)")

    return raw[keep_cols].drop_duplicates(subset="Customer Code")


# ----------------------------------------------------------------------
# REPORT BUILDER
# ----------------------------------------------------------------------
def build_report(sales: pd.DataFrame, customers: pd.DataFrame) -> pd.DataFrame:
    sales = sales.merge(customers, on="Customer Code", how="left")

    # Prefer the name from the sales log; fall back to master list name
    if "Customer Name (Master)" in sales.columns:
        sales["Customer Name"] = sales["Customer Name"].fillna(sales["Customer Name (Master)"])

    # Only consider Feed items (Item No. starts with FEED), excluding returns
    # (rows where Quantity is negative)
    feed_sales = sales[
        sales["Item No."].str.upper().str.startswith(FEED_PREFIX) & (sales["Quantity"] > 0)
    ].copy()

    # Last feed purchase date per customer
    last_feed = feed_sales.groupby("Customer Code")["Date"].max().rename("Last Feed Purchase Date")

    # Static per-customer info (name, zone) — take the latest non-null row
    info = (
        feed_sales.sort_values("Date")
        .groupby("Customer Code")
        .agg({"Customer Name": "last", "Zone": "last"})
    )

    report = info.join(last_feed).reset_index()

    # Due date last Purchase = number of days between today and Last Feed Purchase Date
    today = pd.Timestamp.now().normalize()
    report["Due date last Purchase"] = (today - report["Last Feed Purchase Date"]).dt.days

    # No "Remarks" column exists in the source sheet — kept empty
    report["Remarks"] = ""

    # Last Order = all Item Description + Quantity pairs on that customer's
    # last feed purchase date, combined into a single readable string
    merged = feed_sales.merge(
        report[["Customer Code", "Last Feed Purchase Date"]], on="Customer Code", how="inner"
    )
    same_day = merged[merged["Date"] == merged["Last Feed Purchase Date"]]

    def combine_items(rows: pd.DataFrame) -> str:
        parts = [f"{desc} ({qty:g})" for desc, qty in zip(rows["Item Description"], rows["Quantity"])]
        return ", ".join(parts)

    last_order = same_day.groupby("Customer Code").apply(combine_items).rename("Last Order")

    report = report.merge(last_order, on="Customer Code", how="left")

    report["Last Feed Purchase Date"] = report["Last Feed Purchase Date"].dt.strftime("%Y-%m-%d")

    # Zone kept in the dataframe (used for filtering) but not shown in the final table
    report = report[
        ["Customer Code", "Customer Name", "Zone", "Last Feed Purchase Date",
         "Due date last Purchase", "Remarks", "Last Order"]
    ]
    return report.sort_values("Customer Code").reset_index(drop=True)


def get_feed_sales(sales: pd.DataFrame, customers: pd.DataFrame) -> pd.DataFrame:
    """
    Same feed-only, no-returns filter used by build_report, exposed standalone
    so the Item Wise Custom Feed Purchased Report can reuse it without
    touching build_report's own logic.
    """
    sales = sales.merge(customers, on="Customer Code", how="left")
    if "Customer Name (Master)" in sales.columns:
        sales["Customer Name"] = sales["Customer Name"].fillna(sales["Customer Name (Master)"])
    feed_sales = sales[
        sales["Item No."].str.upper().str.startswith(FEED_PREFIX) & (sales["Quantity"] > 0)
    ].copy()
    return feed_sales


# ----------------------------------------------------------------------
# EXCEL EXPORT — one sheet, zone blocks stacked one after another
# ----------------------------------------------------------------------
def build_excel(report_df: pd.DataFrame, zones_in_order: list, display_cols: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Feed Purchase Report"

    zone_font = Font(bold=True, size=13)
    header_font = Font(bold=True)
    due_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    due_col_idx = display_cols.index("Due date last Purchase") + 1

    row = 1
    for zone in zones_in_order:
        zone_table = report_df[report_df["Zone"] == zone][display_cols]

        # Zone name row (no border/fill — just a section title)
        ws.cell(row=row, column=1, value=zone).font = zone_font
        row += 1

        # Header row
        for col_idx, col_name in enumerate(display_cols, start=1):
            cell = ws.cell(row=row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            if col_idx == due_col_idx:
                cell.fill = yellow_fill
        row += 1

        # Data rows
        for _, data_row in zone_table.iterrows():
            for col_idx, col_name in enumerate(display_cols, start=1):
                cell = ws.cell(row=row, column=col_idx, value=data_row[col_name])
                cell.alignment = center_align
                cell.border = border
                if col_idx == due_col_idx:
                    cell.fill = yellow_fill
                    cell.font = due_font
            row += 1

        # Blank spacer row before the next zone
        row += 1

    for col_idx, col_name in enumerate(display_cols, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(18, len(col_name) + 2)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ----------------------------------------------------------------------
# EXCEL EXPORT — flat table, used by the Item Wise Custom Feed Purchased Report
# ----------------------------------------------------------------------
def build_item_excel(df: pd.DataFrame, item_desc: str, date_range_str: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Item Wise Report"

    title_font = Font(bold=True, size=13)
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: selected Item Description, Row 2: selected Date Range
    ws.cell(row=1, column=1, value=f"Item Description: {item_desc}").font = title_font
    ws.cell(row=2, column=1, value=f"Date Range: {date_range_str}").font = title_font

    cols = list(df.columns)
    header_row = 3
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    for r_idx, (_, data_row) in enumerate(df.iterrows(), start=header_row + 1):
        for col_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=data_row[col_name])
            cell.alignment = center_align
            cell.border = border

    for col_idx, col_name in enumerate(cols, start=1):
        ws.column_dimensions[ws.cell(row=header_row, column=col_idx).column_letter].width = max(18, len(col_name) + 2)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ----------------------------------------------------------------------
# UI
# ----------------------------------------------------------------------
report_choice = st.radio(
    "Select Report",
    ["Last Feed Purchase Date Report", "Item Wise Custom Feed Purchased Report"],
)

st.title(f"📦 {report_choice}")

try:
    sales_df = load_sales_data(SHEET_ID, DEFAULT_GID)
    customers_df = load_customer_master(CUSTOMER_LIST_PATH)
except Exception as e:
    st.error(f"Error loading data: {e}")
    st.stop()

if report_choice == "Last Feed Purchase Date Report":
    report_df = build_report(sales_df, customers_df)

    # Zone-wise slicer — table only appears after a zone is picked
    zones = sorted([z for z in report_df["Zone"].dropna().unique() if z and z != "nan"])

    st.subheader("🌍 Select Zone")
    selected_zones = st.multiselect("Zone", options=zones, placeholder="Choose one or more zones...")

    if not selected_zones:
        st.info("👆 Select at least one zone above to display the report.")
    else:
        filtered = report_df[report_df["Zone"].isin(selected_zones)]

        DISPLAY_COLS = ["Customer Code", "Customer Name", "Last Feed Purchase Date",
                        "Due date last Purchase", "Remarks", "Last Order"]
        table = filtered[DISPLAY_COLS]

        st.dataframe(
            table,
            use_container_width=True,
            hide_index=True,
        )

        excel_bytes = build_excel(report_df, selected_zones, DISPLAY_COLS)
        st.download_button(
            "⬇️ Download Excel Report",
            data=excel_bytes,
            file_name="feed_purchase_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

else:  # Item Wise Custom Feed Purchased Report
    feed_sales_all = get_feed_sales(sales_df, customers_df)

    item_options = sorted(feed_sales_all["Item Description"].dropna().unique())
    min_date = feed_sales_all["Date"].min()
    max_date = feed_sales_all["Date"].max()

    st.subheader("🔍 Select Item & Date Range")
    selected_item = st.selectbox("Item Description", options=item_options)
    date_range = st.date_input(
        "Date Range",
        value=(min_date.date(), max_date.date()),
        min_value=min_date.date(),
        max_value=max_date.date(),
    )

    if isinstance(date_range, tuple) and len(date_range) == 2:
        start_date, end_date = date_range
        mask = (
            (feed_sales_all["Item Description"] == selected_item)
            & (feed_sales_all["Date"] >= pd.Timestamp(start_date))
            & (feed_sales_all["Date"] <= pd.Timestamp(end_date))
        )
        item_table = feed_sales_all.loc[
            mask, ["Date", "Item Description", "Customer Code", "Customer Name", "Quantity"]
        ].copy()
        item_table = item_table.sort_values("Date")
        item_table["Date"] = item_table["Date"].dt.strftime("%Y-%m-%d")
        item_table = item_table.reset_index(drop=True)

        st.dataframe(
            item_table,
            use_container_width=True,
            hide_index=True,
        )

        item_excel_bytes = build_item_excel(
            item_table,
            selected_item,
            f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
        )
        st.download_button(
            "⬇️ Download Excel Report",
            data=item_excel_bytes,
            file_name="item_wise_feed_purchase_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.info("👆 Select a start and end date to display the report.")
